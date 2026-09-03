"""CDE pupil membership (data/raw/cde/*.xlsx) rows for Boulder Valley Re 2 (0480), school codes 0652 Bear Creek
and 5838 Mesa. Output data/clean/cde_mesa_bearcreek.csv and a CDE-vs-BVSD comparison
data/clean/verification_cde_vs_bvsd.csv. School codes confirmed against the 'School Name' column in every file."""
import openpyxl, glob, re
import pandas as pd
rows = []
for f in sorted(glob.glob("data/raw/cde/*_membership_grade_by_school.xlsx")):
    sy = re.search(r"(\d{4}-\d{2})", f).group(1)
    ws = openpyxl.load_workbook(f, read_only=True).worksheets[0]
    hdr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] in ("County Code", "Organization Code"):
            hdr = list(row); continue
        if hdr and row and any(c in ("0652", "5838") for c in row if isinstance(c, str)):
            d = dict(zip(hdr, row))
            if str(d.get("Distr Code", d.get("Organization Code"))) != "0480": continue
            code = d.get("Sch Code", d.get("School Code")); name = d["School Name"]
            r = dict(school={"0652": "Bear Creek", "5838": "Mesa"}[code], cde_school_code=code, cde_school_name=name, school_year=sy,
                     PK=d["Pre-K"], K_half=d["Half-Day K"], K_full=d["Full-Day K"], G1=d["1st"], G2=d["2nd"], G3=d["3rd"], G4=d["4th"], G5=d["5th"],
                     total_pk12=[v for v in row if isinstance(v, (int, float))][-1], _source=f, _row=i, _sheet=ws.title)
            rows.append(r)
# 2025-26 workbook: 'Grade' sheet
f = "data/raw/cde/2025-26_pupil_membership_school_level.xlsx"
wb = openpyxl.load_workbook(f, read_only=True)
ws = wb["Grade"]; hdr = None
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row and row[0] == "Organization Code": hdr = [str(c).replace("\n", " ").strip() for c in row]; continue
    if hdr and row and str(row[0]) == "0480" and str(row[2]) in ("0652", "5838"):
        d = dict(zip(hdr, row))
        keys = list(d.keys()); print("2025-26 Grade sheet columns:", keys)
        rows.append(dict(school={"0652": "Bear Creek", "5838": "Mesa"}[str(row[2])], cde_school_code=str(row[2]), cde_school_name=row[3], school_year="2025-26",
                         PK=d.get("Pre-K", d.get("PK")), K_half=d.get("Half-Day K", d.get("Half-Day Kinder")), K_full=d.get("Full-Day K", d.get("Full-Day Kinder")), G1=d.get("1st"), G2=d.get("2nd"), G3=d.get("3rd"), G4=d.get("4th"), G5=d.get("5th"),
                         total_pk12=d.get("PK-12 Count"), _source=f, _row=i, _sheet="Grade"))
cde = pd.DataFrame(rows)
for col in ["PK", "K_half", "K_full", "G1", "G2", "G3", "G4", "G5", "total_pk12"]:
    cde[col] = pd.to_numeric(cde[col], errors="coerce")  # CDE prints "-" for suppressed/zero cells
cde["K_5"] = cde[["K_half", "K_full", "G1", "G2", "G3", "G4", "G5"]].sum(axis=1)
cde.to_csv("data/clean/cde_mesa_bearcreek.csv", index=False)
bv = pd.read_csv("data/clean/bvsd_pupil_count_mesa_bearcreek.csv"); bv = bv[bv.file_family == "headcount"]
m = cde.merge(bv[["school", "school_year", "funded_headcount", "K", "G1", "G2", "G3", "G4", "G5"]], on=["school", "school_year"], suffixes=("_cde", "_bvsd"))
m["cde_K5_minus_bvsd"] = m.K_5 - m.funded_headcount
m["cde_PK12_minus_bvsd"] = m.total_pk12 - m.funded_headcount
grade_diff = {g: (m[f"{g}_cde"] - m[f"{g}_bvsd"]) for g in ["G1", "G2", "G3", "G4", "G5"]}
m["K_diff"] = (m.K_half + m.K_full) - m.K
for g, s in grade_diff.items(): m[f"{g}_diff"] = s
out = m[["school", "school_year", "total_pk12", "PK", "K_5", "funded_headcount", "cde_K5_minus_bvsd", "cde_PK12_minus_bvsd", "K_diff", "G1_diff", "G2_diff", "G3_diff", "G4_diff", "G5_diff"]]
out.to_csv("data/clean/verification_cde_vs_bvsd.csv", index=False)
print(out.to_string(index=False))
# FRL sheets 2025-26
for sh in ["FRL_K12", "FRL_PK12", "PK12_MembershipTrends"]:
    ws = wb[sh]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "Organization Code": print(sh, "cols:", [str(c).replace("\n"," ") for c in row])
        if row and str(row[0]) == "0480" and str(row[2]) in ("0652", "5838"): print(sh, row)
